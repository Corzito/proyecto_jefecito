import io
import openpyxl
from datetime import datetime, date as date_type
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.http import JsonResponse, HttpResponse
from django.core.management import call_command
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Colaborador
from .forms import ColaboradorForm


def lista_colaboradores(request):
    colaboradores = Colaborador.objects.all()
    hoy = timezone.now().date()

    q = request.GET.get('q', '').strip()
    empresa_filtro = request.GET.get('empresa', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    orden = request.GET.get('orden', 'fecha_ingreso')

    if q:
        from django.db.models import Q
        colaboradores = colaboradores.filter(
            Q(nombres__icontains=q) | Q(cedula__icontains=q)
        )
    if empresa_filtro:
        colaboradores = colaboradores.filter(empresa=empresa_filtro)
    if fecha_desde:
        try:
            colaboradores = colaboradores.filter(fecha_ingreso__gte=fecha_desde)
        except Exception:
            pass
    if fecha_hasta:
        try:
            colaboradores = colaboradores.filter(fecha_ingreso__lte=fecha_hasta)
        except Exception:
            pass

    if orden == 'nombres':
        colaboradores = colaboradores.order_by('nombres')
    elif orden == 'empresa':
        colaboradores = colaboradores.order_by('empresa', 'nombres')
    else:
        colaboradores = colaboradores.order_by('fecha_ingreso', 'nombres')

    data = []
    alertas_pendientes = 0
    en_seguimiento = 0
    completados_count = 0

    for col in colaboradores:
        dias = (hoy - col.fecha_ingreso).days
        estado = col.estado_periodo()
        if estado in ('alerta_30', 'alerta_50'):
            alertas_pendientes += 1
        if estado in ('en_seguimiento', 'alerta_50'):
            en_seguimiento += 1
        if estado == 'completado':
            completados_count += 1
        data.append({
            'obj': col,
            'dias': dias,
            'estado': estado,
            'dias_para_30': max(0, 30 - dias),
            'dias_para_50': max(0, 50 - dias),
        })

    return render(request, '_periodo_de_prueba/lista.html', {
        'colaboradores': data,
        'alertas_pendientes': alertas_pendientes,
        'en_seguimiento': en_seguimiento,
        'completados_count': completados_count,
        'hoy': hoy,
        'q': q,
        'empresa_filtro': empresa_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'orden': orden,
    })


def agregar_colaborador(request):
    if request.method == 'POST':
        form = ColaboradorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Colaborador registrado correctamente.')
            return redirect('periodo:lista')
    else:
        form = ColaboradorForm()
    return render(request, '_periodo_de_prueba/form.html', {'form': form, 'titulo': 'Registrar Colaborador'})


def editar_colaborador(request, pk):
    colaborador = get_object_or_404(Colaborador, pk=pk)
    if request.method == 'POST':
        form = ColaboradorForm(request.POST, instance=colaborador)
        if form.is_valid():
            form.save()
            messages.success(request, 'Colaborador actualizado.')
            return redirect('periodo:lista')
    else:
        form = ColaboradorForm(instance=colaborador)
    return render(request, '_periodo_de_prueba/form.html', {'form': form, 'titulo': 'Editar Colaborador'})


def eliminar_colaborador(request, pk):
    colaborador = get_object_or_404(Colaborador, pk=pk)
    if request.method == 'POST':
        colaborador.delete()
        messages.success(request, 'Colaborador eliminado.')
        return redirect('periodo:lista')
    return render(request, '_periodo_de_prueba/confirmar_eliminar.html', {'colaborador': colaborador})


def marcar_evaluacion(request, pk, tipo):
    colaborador = get_object_or_404(Colaborador, pk=pk)
    if tipo == '30':
        colaborador.evaluacion_30_completada = True
    elif tipo == '50':
        colaborador.evaluacion_50_completada = True
    colaborador.save()
    messages.success(request, f'Evaluación de {tipo} días marcada como completada.')
    return redirect('periodo:lista')


def marcar_resultado(request, pk):
    colaborador = get_object_or_404(Colaborador, pk=pk)
    if request.method == 'POST':
        resultado = request.POST.get('resultado', 'pendiente')
        observaciones = request.POST.get('observaciones', '')
        colaborador.resultado_periodo = resultado
        colaborador.observaciones = observaciones
        colaborador.save()
        messages.success(request, f'Resultado de {colaborador.nombres} actualizado.')
        return redirect('periodo:completados')
    return render(request, '_periodo_de_prueba/marcar_resultado.html', {
        'colaborador': colaborador,
    })


def primer_periodo(request):
    hoy = timezone.now().date()
    q = request.GET.get('q', '').strip()
    colaboradores = Colaborador.objects.all()

    if q:
        from django.db.models import Q
        colaboradores = colaboradores.filter(
            Q(nombres__icontains=q) | Q(cedula__icontains=q)
        )

    data = []
    for col in colaboradores:
        dias = (hoy - col.fecha_ingreso).days
        if dias < 30:
            estado = col.estado_periodo()
            data.append({
                'obj': col,
                'dias': dias,
                'estado': estado,
                'dias_para_30': max(0, 30 - dias),
            })

    return render(request, '_periodo_de_prueba/primer_periodo.html', {
        'colaboradores': data,
        'hoy': hoy,
        'q': q,
        'total': len(data),
    })


def segundo_periodo(request):
    hoy = timezone.now().date()
    q = request.GET.get('q', '').strip()
    colaboradores = Colaborador.objects.all()

    if q:
        from django.db.models import Q
        colaboradores = colaboradores.filter(
            Q(nombres__icontains=q) | Q(cedula__icontains=q)
        )

    data = []
    for col in colaboradores:
        dias = (hoy - col.fecha_ingreso).days
        if 30 <= dias < 50:
            estado = col.estado_periodo()
            data.append({
                'obj': col,
                'dias': dias,
                'estado': estado,
                'dias_para_50': max(0, 50 - dias),
            })

    return render(request, '_periodo_de_prueba/segundo_periodo.html', {
        'colaboradores': data,
        'hoy': hoy,
        'q': q,
        'total': len(data),
    })


def completados(request):
    hoy = timezone.now().date()
    q = request.GET.get('q', '').strip()
    resultado_filtro = request.GET.get('resultado', '').strip()
    colaboradores = Colaborador.objects.all()

    if q:
        from django.db.models import Q
        colaboradores = colaboradores.filter(
            Q(nombres__icontains=q) | Q(cedula__icontains=q)
        )
    if resultado_filtro:
        colaboradores = colaboradores.filter(resultado_periodo=resultado_filtro)

    data = []
    for col in colaboradores:
        dias = (hoy - col.fecha_ingreso).days
        if dias >= 50:
            data.append({
                'obj': col,
                'dias': dias,
            })

    return render(request, '_periodo_de_prueba/completados.html', {
        'colaboradores': data,
        'hoy': hoy,
        'q': q,
        'resultado_filtro': resultado_filtro,
        'total': len(data),
    })


def importar_excel(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        if not archivo:
            messages.error(request, 'Selecciona un archivo Excel.')
            return redirect('periodo:lista')
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, 'El archivo debe ser .xlsx')
            return redirect('periodo:lista')
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception:
            messages.error(request, 'No se pudo leer el archivo.')
            return redirect('periodo:lista')

        exitosos = 0
        errores = []
        duplicados = 0

        encabezados = {}
        for cell in ws[1]:
            if cell.value:
                encabezados[str(cell.value).upper().strip()] = cell.column - 1

        mapa = {
            "CÉDULA NO": "cedula", "CEDULA NO": "cedula", "CEDULA": "cedula",
            "NOMBRES COMPLETOS": "nombres", "NOMBRES": "nombres",
            "CARGO": "cargo",
            "JEFE INMEDIATO": "jefe_inmediato",
            "EMPRESA": "empresa",
            "NO CELULAR": "celular", "CELULAR": "celular",
            "FECHA INGRESO (AAAA-MM-DD)": "fecha_ingreso", "FECHA INGRESO": "fecha_ingreso",
        }

        indices = {}
        for enc_label, campo in mapa.items():
            if enc_label in encabezados and campo not in indices:
                indices[campo] = encabezados[enc_label]

        campos_requeridos = ["cedula", "nombres", "cargo", "jefe_inmediato", "empresa", "celular", "fecha_ingreso"]
        faltantes = [c for c in campos_requeridos if c not in indices]
        if faltantes:
            messages.error(request, f"Columnas faltantes: {', '.join(faltantes)}. Usa la plantilla oficial.")
            return redirect('periodo:lista')

        EMPRESAS_VALIDAS = {"CARBOINSA", "INCARSA", "UNIMINAS", "MILPA"}

        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None or str(v).strip() == "" for v in fila):
                continue

            def get(campo):
                idx = indices.get(campo)
                if idx is None:
                    return ""
                val = fila[idx]
                return str(val).strip() if val is not None else ""

            cedula    = get("cedula")
            nombres   = get("nombres")
            cargo     = get("cargo")
            jefe      = get("jefe_inmediato")
            empresa   = get("empresa").upper()
            celular   = get("celular")
            fecha_str = get("fecha_ingreso")

            if not cedula or not nombres:
                errores.append(f'Fila {fila_num}: cédula o nombre vacío.')
                duplicados += 1
                continue

            if empresa not in EMPRESAS_VALIDAS:
                errores.append(f'Fila {fila_num} ({nombres}): empresa "{empresa}" no válida.')
                duplicados += 1
                continue

            try:
                val_fecha = fila[indices["fecha_ingreso"]]
                # ✅ FIX: normalizar datetime a date si viene como datetime del Excel
                if isinstance(val_fecha, datetime):
                    fecha_ingreso = val_fecha.date()
                elif isinstance(val_fecha, date_type):
                    fecha_ingreso = val_fecha
                else:
                    fecha_ingreso = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                errores.append(f'Fila {fila_num} ({nombres}): fecha "{fecha_str}" inválida.')
                duplicados += 1
                continue

            if Colaborador.objects.filter(cedula=cedula).exists():
                duplicados += 1
                continue

            dias = (date_type.today() - fecha_ingreso).days
            alerta_30 = dias >= 23
            alerta_50 = dias >= 43

            Colaborador.objects.create(
                cedula=cedula,
                nombres=nombres,
                cargo=cargo,
                jefe_inmediato=jefe,
                empresa=empresa,
                celular=celular,
                fecha_ingreso=fecha_ingreso,
                alerta_30_enviada=alerta_30,
                alerta_50_enviada=alerta_50,
            )
            exitosos += 1

        if exitosos:
            messages.success(request, f'✅ {exitosos} colaborador(es) importado(s) correctamente.')
        if duplicados:
            messages.warning(request, f'⚠️ {duplicados} fila(s) omitida(s).')
        for err in errores[:5]:
            messages.error(request, err)

        return redirect('periodo:lista')

    return redirect('periodo:lista')


def descargar_plantilla(request):
    rojo       = "E32822"
    blanco     = "FFFFFF"
    gris_claro = "F2F2F2"

    header_font  = Font(bold=True, color=blanco, size=11)
    header_fill  = PatternFill("solid", fgColor=rojo)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    data_align = Alignment(horizontal="left", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"

    columnas = [
        ("Cédula No",                   20),
        ("Nombres Completos",            35),
        ("Cargo",                        35),
        ("Jefe Inmediato",               28),
        ("Empresa",                      18),
        ("No Celular",                   18),
        ("Fecha Ingreso (AAAA-MM-DD)",   26),
    ]

    for col_idx, (label, width) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32

    data_fill_par   = PatternFill("solid", fgColor=gris_claro)
    data_fill_impar = PatternFill("solid", fgColor=blanco)

    for row in range(2, 52):
        fill = data_fill_par if row % 2 == 0 else data_fill_impar
        for col in range(1, len(columnas) + 1):
            cell = ws.cell(row=row, column=col, value=None)
            cell.border    = thin_border
            cell.fill      = fill
            cell.alignment = data_align
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A2"

    wi = wb.create_sheet("📋 Instrucciones")
    wi.column_dimensions["A"].width = 32
    wi.column_dimensions["B"].width = 55

    titulo_font = Font(bold=True, color=blanco, size=13)
    titulo_fill = PatternFill("solid", fgColor=rojo)
    ok_fill     = PatternFill("solid", fgColor="E2EFDA")
    ok_font     = Font(color="375623", size=10)

    wi.merge_cells("A1:B1")
    c = wi.cell(row=1, column=1, value="📋  INSTRUCCIONES DE DILIGENCIAMIENTO")
    c.font      = titulo_font
    c.fill      = titulo_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border
    wi.row_dimensions[1].height = 30

    instrucciones = [
        ("Cédula No",                  "Número sin puntos ni espacios.",          "1002368124"),
        ("Nombres Completos",          "Nombre y apellidos en mayúsculas.",        "ALDO FLECHAS ALVAREZ"),
        ("Cargo",                      "Cargo o rol del colaborador.",             "APRENDIZ SENA"),
        ("Jefe Inmediato",             "Nombre del jefe inmediato.",               "ING. JAVIER MOJICA"),
        ("Empresa",                    "CARBOINSA | INCARSA | UNIMINAS | MILPA",   "INCARSA"),
        ("No Celular",                 "Sin espacios ni guiones.",                 "3137774696"),
        ("Fecha Ingreso (AAAA-MM-DD)", "Formato AÑO-MES-DÍA.",                    "2025-01-10"),
    ]

    fila_act = 2
    for campo, desc, ejemplo_val in instrucciones:
        c = wi.cell(row=fila_act, column=1, value=campo)
        c.font      = Font(size=10, bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = thin_border
        wi.row_dimensions[fila_act].height = 36

        c = wi.cell(row=fila_act, column=2, value=desc)
        c.font      = Font(size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = thin_border

        fila_act += 1
        c = wi.cell(row=fila_act, column=1, value="Ejemplo →")
        c.font = ok_font; c.fill = ok_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border

        c = wi.cell(row=fila_act, column=2, value=ejemplo_val)
        c.font = ok_font; c.fill = ok_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border
        wi.row_dimensions[fila_act].height = 18
        fila_act += 1

    we = wb.create_sheet("Empresas válidas")
    we.column_dimensions["A"].width = 35
    c = we.cell(row=1, column=1, value="Valores aceptados en la columna EMPRESA")
    c.font      = Font(bold=True, color=blanco, size=11)
    c.fill      = PatternFill("solid", fgColor=rojo)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border
    we.row_dimensions[1].height = 28
    for i, emp in enumerate(["CARBOINSA", "INCARSA", "UNIMINAS", "MILPA"], start=2):
        c = we.cell(row=i, column=1, value=emp)
        c.font      = Font(size=11, bold=True)
        c.fill      = PatternFill("solid", fgColor="E2EFDA")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border
        we.row_dimensions[i].height = 22

    wb.active = ws
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_colaboradores.xlsx"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# REPORTE EXCEL — filtro por fecha + selección de colaboradores
# ─────────────────────────────────────────────────────────────────────────────

def reporte_excel(request):
    """
    GET  → muestra la página con filtros y lista de colaboradores seleccionables.
    POST → genera y descarga el Excel con los colaboradores seleccionados.
    """
    hoy = timezone.now().date()

    # Parámetros de filtro (presentes en GET y en POST)
    fecha_desde    = request.POST.get('fecha_desde', request.GET.get('fecha_desde', '')).strip()
    fecha_hasta    = request.POST.get('fecha_hasta', request.GET.get('fecha_hasta', '')).strip()
    empresa_filtro = request.POST.get('empresa', request.GET.get('empresa', '')).strip()

    colaboradores = Colaborador.objects.all().order_by('empresa', 'nombres')

    if fecha_desde:
        try:
            colaboradores = colaboradores.filter(fecha_ingreso__gte=fecha_desde)
        except Exception:
            pass
    if fecha_hasta:
        try:
            colaboradores = colaboradores.filter(fecha_ingreso__lte=fecha_hasta)
        except Exception:
            pass
    if empresa_filtro:
        colaboradores = colaboradores.filter(empresa=empresa_filtro)

    # Construir data con días calculados
    data = []
    for col in colaboradores:
        dias = (hoy - col.fecha_ingreso).days
        estado = col.estado_periodo()
        etiqueta_estado = {
            'en_seguimiento': 'En seguimiento',
            'alerta_30':      'Alerta 30 días',
            'alerta_50':      'Alerta 50 días',
            'completado':     'Completado',
        }.get(estado, estado)
        data.append({
            'obj': col,
            'dias': dias,
            'estado': etiqueta_estado,
        })

    # ── POST: generar el Excel ──────────────────────────────────────────────
    if request.method == 'POST':
        ids_seleccionados = request.POST.getlist('colaboradores_ids')
        if not ids_seleccionados:
            messages.warning(request, 'Selecciona al menos un colaborador para generar el reporte.')
            return render(request, '_periodo_de_prueba/reporte_excel.html', {
                'data': data,
                'hoy': hoy,
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
                'empresa_filtro': empresa_filtro,
            })

        seleccionados = [d for d in data if str(d['obj'].pk) in ids_seleccionados]

        rojo   = "E32822"
        blanco = "FFFFFF"
        gris   = "F2F2F2"

        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        header_font  = Font(bold=True, color=blanco, size=10)
        header_fill  = PatternFill("solid", fgColor=rojo)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center       = Alignment(horizontal="center", vertical="center")
        left         = Alignment(horizontal="left",   vertical="center")

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Periodo de Prueba"

        # ── Título del reporte ──────────────────────────────────────────────
        ws.merge_cells("A1:K1")
        titulo_cell = ws["A1"]
        titulo_cell.value     = "REPORTE PERIODO DE PRUEBA — GRUPO EMPRESARIAL"
        titulo_cell.font      = Font(bold=True, color=blanco, size=13)
        titulo_cell.fill      = PatternFill("solid", fgColor=rojo)
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        titulo_cell.border    = thin
        ws.row_dimensions[1].height = 28

        # ── Subtítulo con rango de fechas ───────────────────────────────────
        ws.merge_cells("A2:K2")
        rango_txt = ""
        if fecha_desde and fecha_hasta:
            rango_txt = f"Ingreso entre {fecha_desde} y {fecha_hasta}"
        elif fecha_desde:
            rango_txt = f"Ingreso desde {fecha_desde}"
        elif fecha_hasta:
            rango_txt = f"Ingreso hasta {fecha_hasta}"
        if empresa_filtro:
            rango_txt = (rango_txt + f" | Empresa: {empresa_filtro}").strip(" | ")
        sub_cell = ws["A2"]
        sub_cell.value     = rango_txt or "Todos los colaboradores seleccionados"
        sub_cell.font      = Font(italic=True, size=10, color="555555")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        sub_cell.border    = thin
        ws.row_dimensions[2].height = 18

        # Fila en blanco
        ws.row_dimensions[3].height = 6

        # ── Encabezados de columna ──────────────────────────────────────────
        encabezados = [
            ("N°",                        5),
            ("Cédula",                    15),
            ("Nombres Completos",         35),
            ("Cargo",                     28),
            ("Jefe Inmediato",            25),
            ("Empresa",                   14),
            ("Celular",                   14),
            ("Fecha Ingreso",             16),
            ("Días Transcurridos",        10),
            ("Estado",                    18),
            ("Resultado / Observaciones", 30),
        ]

        for col_idx, (label, width) in enumerate(encabezados, start=1):
            cell = ws.cell(row=4, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[4].height = 30

        # ── Filas de datos ──────────────────────────────────────────────────
        fill_par   = PatternFill("solid", fgColor=gris)
        fill_impar = PatternFill("solid", fgColor=blanco)

        # Color de estado
        color_estado = {
            'En seguimiento': ("FFF2CC", "7F6000"),  # amarillo
            'Alerta 30 días': ("FCE4D6", "833C0B"),  # naranja
            'Alerta 50 días': ("F4CCCC", "990000"),  # rojo claro
            'Completado':     ("E2EFDA", "375623"),  # verde
        }

        resultado_labels = {
            'aprobado':    '✅ Aprobado',
            'no_aprobado': '❌ No aprobado',
            'pendiente':   '⏳ Pendiente',
        }

        for fila_idx, item in enumerate(seleccionados, start=1):
            col = item['obj']
            fila_excel = fila_idx + 4
            fill = fill_par if fila_idx % 2 == 0 else fill_impar
            resultado_txt = resultado_labels.get(col.resultado_periodo, col.resultado_periodo or '')
            if col.observaciones:
                resultado_txt += f" — {col.observaciones}"

            valores = [
                fila_idx,
                col.cedula,
                col.nombres,
                col.cargo,
                col.jefe_inmediato,
                col.empresa,
                col.celular,
                col.fecha_ingreso.strftime("%Y-%m-%d"),
                item['dias'],
                item['estado'],
                resultado_txt,
            ]

            for col_idx, valor in enumerate(valores, start=1):
                cell = ws.cell(row=fila_excel, column=col_idx, value=valor)
                cell.border    = thin
                cell.alignment = center if col_idx in (1, 9) else left
                # Estado con color
                if col_idx == 10:
                    bg, fg = color_estado.get(item['estado'], (gris, "000000"))
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(bold=True, color=fg, size=10)
                else:
                    cell.fill = fill
                    cell.font = Font(size=10)
            ws.row_dimensions[fila_excel].height = 18

        # ── Fila de total ───────────────────────────────────────────────────
        fila_total = len(seleccionados) + 5
        ws.merge_cells(f"A{fila_total}:H{fila_total}")
        tc = ws.cell(row=fila_total, column=1, value=f"Total de colaboradores en el reporte: {len(seleccionados)}")
        tc.font      = Font(bold=True, size=10, color=blanco)
        tc.fill      = PatternFill("solid", fgColor=rojo)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.border    = thin
        for col_idx in range(2, 12):
            c = ws.cell(row=fila_total, column=col_idx)
            c.border = thin
            if col_idx <= 8:
                c.fill = PatternFill("solid", fgColor=rojo)
        ws.row_dimensions[fila_total].height = 20

        ws.freeze_panes = "A5"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f"reporte_periodo_prueba_{hoy.strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
        return response

    # ── GET: mostrar la página ──────────────────────────────────────────────
    return render(request, '_periodo_de_prueba/reporte_excel.html', {
        'data': data,
        'hoy': hoy,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'empresa_filtro': empresa_filtro,
    })


def ejecutar_alertas(request):
    token = request.GET.get('token', '')
    if token != 'incarsa2026seguro':
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        call_command('enviar_alertas_periodo')
        return JsonResponse({'status': 'ok', 'mensaje': 'Alertas procesadas correctamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)


def ping(request):
    return JsonResponse({'status': 'ok'})